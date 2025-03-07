from flask import Flask, request, jsonify, send_file, render_template
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
import io
import os
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side
from flask_cors import CORS 
from dotenv import load_dotenv
import traceback

load_dotenv()

app = Flask(__name__, static_folder='./static', template_folder='./template')
CORS(app)

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')

app.config['SQLALCHEMY_DATABASE_URI'] = f"mysql+pymysql://{os.getenv('DB_USER')}:{os.getenv('DB_PASSWORD')}@{os.getenv('DB_HOST')}/{os.getenv('DB_NAME')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

print("Conexão efetuada com sucesso na porta http://127.0.0.1:5000")

db = SQLAlchemy(app)

@app.route('/')
def home():
    page = render_template('index.html')
    return page

@app.route('/csv', methods=['POST'])
def gerar_csv():
    try:
        registro = request.get_json()
        data_inicio = registro.get('data-inicio')
        data_fim = registro.get('data-fim')

        if not data_inicio or not data_fim:
            print("Data Início ou Data Fim não inseridos.")
            return jsonify({'erro': 'Data Início ou Data Fim não inseridos.'}), 404
        

        query = """
                SELECT
                    ACTION_CARD_ID BS_CARD_ID,
                    CARD_CREATION_DATE,
                    CARD_NAME,
                    CARD_ID_SHORT,
                    MAX(ACTION_DATE) AS CONCLUSION_DATE
                FROM
                    ACTIONS
                INNER JOIN
                    CARDS ON CARD_ID = ACTION_CARD_ID
                WHERE
                    ACTION_LIST_AFTER IN (
                        SELECT RULES_TRELLO_OBJECT_ID
                        FROM RULES
                        WHERE RULES_KEY = 'doneList'
                        AND RULES_ACTIVE = 1
                    )
                    AND ACTION_DATE BETWEEN %s AND %s
                GROUP BY
                    ACTION_CARD_ID

                UNION ALL

                SELECT
                    ACTION_CARD_ID BS_CARD_ID,
                    CARD_CREATION_DATE,
                    CARD_NAME,
                    CARD_ID_SHORT,
                    MAX(ACTION_DATE) AS CONCLUSION_DATE
                FROM
                    ACTIONS
                INNER JOIN
                    CARDS ON CARD_ID = ACTION_CARD_ID
                WHERE
                    ACTION_LIST_AFTER IN (
                        SELECT RULES_TRELLO_OBJECT_ID
                        FROM RULES
                        WHERE RULES_KEY = 'doneList'
                        AND RULES_ACTIVE = 0
                    )
                    AND ACTION_DATE BETWEEN %s AND %s
                GROUP BY
                    ACTION_CARD_ID;
            """

        # query = "SELECT ACTION_CARD_ID BS_CARD_ID, MAX(ACTION_DATE) AS CONCLUSION_DATE FROM ACTIONS WHERE ACTION_LIST_AFTER = '670d1616ad6d3d830c285c41' AND ACTION_DATE BETWEEN %s AND %s GROUP BY  ACTION_CARD_ID"

        print("Data Inicial: " + data_inicio)
        print("Data Final: " + data_fim)

        with db.engine.connect() as conexao:
            df = pd.read_sql(query, conexao, params=(data_inicio, data_fim, data_inicio, data_fim))

            print(df)


            data_cards = requests.get(os.getenv('TRELLO_CARDS_URL'))

            data_lists = requests.get(os.getenv('TRELLO_LISTS_URL'))

            board_cards = data_cards.json()
            data_lists = data_lists.json()

            lists = [{'idList': lst['id'], 'listName': lst['name']} for lst in data_lists]
            lists_df = pd.DataFrame(lists)

            lists_filter = ['670d1616ad6d3d830c285c41', '62388db5b91b032488cea097']

            cards = []
            for card in board_cards:
                if card['labels']:
                    for label in card['labels']:
                        cards.append({
                            'BS_CARD_ID': card['id'],
                            'cardName': card['name'],
                            'cardDesc': card['desc'],
                            'idList': card['idList'],
                            'idShort': card['idShort'],
                            'shortUrl': card['shortUrl'],
                            'idLabel': label['id'],
                            'labelName': label['name'],
                            'labelColor': label['color']
                        })
                else:
                    cards.append({
                        'BS_CARD_ID': card['id'],
                        'cardName': card['name'],
                        'idList': card['idList'],
                        'idShort': card['idShort'],
                        'shortUrl': card['shortUrl'],
                        'idLabel': '',
                        'labelName': '',
                        'labelColor': ''
                    })

            cards_df = pd.DataFrame(cards)
            cards_df = cards_df[cards_df['idList'].isin(lists_filter)]

            merge_df = pd.merge(cards_df, lists_df, on='idList')
            merge_df = pd.merge(merge_df, df, on='BS_CARD_ID')

            done_list_df = merge_df[merge_df['labelColor'] == 'lime']

            
            wb = Workbook()
            ws = wb.active

            
            for r_idx, row in enumerate(dataframe_to_rows(done_list_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            
            alignment = Alignment(horizontal='center', vertical='center')
            border = Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                ws.row_dimensions[row[0].row].height = 30
                for cell in row:
                    ws.column_dimensions[cell.column_letter].width = 30
                    cell.alignment = alignment
                    cell.border = border

            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(output, as_attachment=True, download_name=f"entregas_da_semana_{data_inicio}_a_{data_fim}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print("Erro ao executar requisição: ", str(e))
        print(traceback.format_exc());
        return jsonify({'erro': str(e)}), 500


@app.route('/leadtime', methods=['POST'])
def gerar_leadtime():
    try:
        registro = request.get_json()
        data_inicio = registro.get('data-inicio')
        data_fim = registro.get('data-fim')

        if not data_inicio or not data_fim:
            print("Data Início ou Data Fim não inseridos.")
            return jsonify({'erro': 'Data Início ou Data Fim não inseridos.'}), 404

        query = """
            SELECT ACTION_CARD_ID AS cardId,
                   CARD_CREATION_DATE,
                   MAX(ACTION_DATE) AS CONCLUSION_DATE
            FROM ACTIONS
            INNER JOIN CARDS ON CARD_ID = ACTION_CARD_ID
            WHERE ACTION_LIST_AFTER = '62388db5b91b032488cea097'
            AND ACTION_DATE BETWEEN %s AND %s 
            GROUP BY ACTION_CARD_ID, CARD_CREATION_DATE
        """

        print(f"Data Início: {data_inicio}")
        print(f"Data Fim: {data_fim}")

        with db.engine.connect() as conexao:
            df = pd.read_sql(query, conexao, params=(data_inicio, data_fim))
            print(df)

        data_cards = requests.get(os.getenv('TRELLO_CARDS_URL'))
        data_lists = requests.get(os.getenv('TRELLO_LISTS_URL'))

        board_cards = data_cards.json()
        data_lists = data_lists.json()

        lists = [{'idList': lista['id'], 'listName': lista['name']} for lista in data_lists]
        lists_df = pd.DataFrame(lists)
        lists_filter = ['62388db5b91b032488cea097']

        cards = []
        for card in board_cards:
            if card.get('labels'):
                for label in card['labels']:
                    cards.append({
                        'cardId': card['id'],
                        'cardName': card['name'],
                        'cardDesc': card['desc'],
                        'idList': card['idList'],
                        'idShort': card['idShort'],
                        'shortUrl': card['shortUrl'],
                        'idLabel': label['id'],
                        'labelName': label['name'],
                        'labelColor': label['color']
                    })
            else:
                cards.append({
                    'cardId': card['id'],
                    'cardName': card['name'],
                    'idList': card['idList'],
                    'idShort': card['idShort'],
                    'shortUrl': card['shortUrl'],
                    'idLabel': '',
                    'labelName': '',
                    'labelColor': ''
                })

        cards_df = pd.DataFrame(cards)
        cards_df = cards_df[cards_df['idList'].isin(lists_filter)]

        merge_df = pd.merge(cards_df, lists_df, on='idList')
        merge_df = pd.merge(merge_df, df, on='cardId')

        clients_label_filter = merge_df['idLabel'].isin([
            '63ce80a5ed837d03dc64c23b', 
            '63ce808818d9320640a435d2', 
            '63ce80943c3d45017d481a5b'
        ])

        done_list_df = merge_df[clients_label_filter].copy()
        done_list_df.drop(['cardId', 'idList', 'idLabel', 'labelColor', 'listName'], axis=1, inplace=True)

        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(done_list_df, index=False, header=True):
            ws.append(r)

        alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 30
            for cell in row:
                    ws.column_dimensions[cell.column_letter].width = 30
                    cell.alignment = alignment
                    cell.border = border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, as_attachment=True, download_name=f"leadtime_por_complexidade_{data_inicio}_a_{data_fim}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    except Exception as e:
        print(f"Erro ao executar requisição: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'erro': str(e)}), 500

@app.route('/favicon.ico')
def favicon():
    return "", 204

if __name__ == "__main__":
    app.run(debug=True)

